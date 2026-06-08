import os
import json
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

def count_categories_and_subcats(md_file_path: str):
    """Robust counting for design spec MD files"""
    if not os.path.exists(md_file_path):
        return 0, 0
    with open(md_file_path, 'r', encoding='utf-8') as f:
        content = f.read().lower()
    categories = len([line for line in content.splitlines() if "## category" in line])
    subcats = len([line for line in content.splitlines() if "### sub-category" in line])
    return categories, subcats


def count_yara_rules(yara_file_path: str):
    """Count lines, rules in .yara file"""
    if not os.path.exists(yara_file_path):
        return 0, 0, 0, 0
    with open(yara_file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    lines = len(content.splitlines())
    rules = len([line for line in content.splitlines() if line.strip().startswith("rule ")])
    cats = rules // 3 if rules > 5 else rules
    subcats = rules // 2 if rules > 5 else rules
    return lines, rules, cats, subcats


def count_file_lines(file_path: str):
    """Count number of lines in any file"""
    if not os.path.exists(file_path):
        return 0
    with open(file_path, 'r', encoding='utf-8') as f:
        return len(f.readlines())


def generate_report_summary():
    print("📊 Generating Pipeline Summary Report...")

    output_dir = "summary_reports"
    os.makedirs(output_dir, exist_ok=True)

    data = []

    trimmed_dir = "trimmed_reports"
    input_dir = "input_reports"
    design_dir = "design_docs"
    llm_design_dir = os.path.join("design_docs", "llmgenerated")
    yara_py_dir = "yara_rules"
    yara_llm_dir = os.path.join("yara_rules", "llmspecgen")

    for filename in os.listdir(trimmed_dir):
        if not filename.startswith("trim_") or not filename.endswith(".json"):
            continue

        base_name = filename.replace("trim_", "").replace(".json", "")
        input_name = filename.replace("trim_", "")

        row = {
            "File name": input_name,
            "Trimmed File output metrics": filename,
        }

        # === Input File Line Count (New) ===
        input_file_path = os.path.join(input_dir, input_name)
        row["Input File # of lines"] = count_file_lines(input_file_path)

        # Trimmed metrics
        trim_path = os.path.join(trimmed_dir, filename)
        with open(trim_path, 'r', encoding='utf-8') as f:
            trimmed_data = json.load(f)
        row["Record length - Total lines in trimmed output"] = len(json.dumps(trimmed_data, indent=2).splitlines())

        # Python Generated Design Spec
        py_design_file = f"Reqspec_{base_name}_pygen.md"
        py_design_path = os.path.join(design_dir, py_design_file)
        if os.path.exists(py_design_path):
            cat, sub = count_categories_and_subcats(py_design_path)
            row["Python Generated Design spec"] = py_design_file
            row["Python # of Category"] = cat
            row["Python # of sub-category"] = sub
        else:
            row["Python Generated Design spec"] = "-"
            row["Python # of Category"] = 0
            row["Python # of sub-category"] = 0

        # LLM Generated Design Spec
        llm_design_file = f"Reqspec_LLM_{base_name}.md"
        llm_design_path = os.path.join(llm_design_dir, llm_design_file)
        if os.path.exists(llm_design_path):
            cat, sub = count_categories_and_subcats(llm_design_path)
            row["LLM Generated Design spec"] = llm_design_file
            row["LLM # of Category"] = cat
            row["LLM # of sub-category"] = sub
        else:
            row["LLM Generated Design spec"] = "-"
            row["LLM # of Category"] = 0
            row["LLM # of sub-category"] = 0

        # YARA from Python Spec
        yara_py_file = f"{base_name}_pygen.yara"
        yara_py_path = os.path.join(yara_py_dir, yara_py_file)
        if os.path.exists(yara_py_path):
            lines, rules, cats, subs = count_yara_rules(yara_py_path)
            row["LLM generated Yara scripts from python spec"] = yara_py_file
            row["Py YARA # of lines"] = lines
            row["Py YARA # of rules"] = rules
            row["Py YARA # of Category"] = cats
            row["Py YARA # of Sub-category"] = subs
        else:
            row["LLM generated Yara scripts from python spec"] = "-"
            row["Py YARA # of lines"] = 0
            row["Py YARA # of rules"] = 0
            row["Py YARA # of Category"] = 0
            row["Py YARA # of Sub-category"] = 0

        # YARA from LLM Spec
        yara_llm_file = f"LLM_{base_name}.yara"
        yara_llm_path = os.path.join(yara_llm_dir, yara_llm_file)
        if os.path.exists(yara_llm_path):
            lines, rules, cats, subs = count_yara_rules(yara_llm_path)
            row["LLM Generated Yara Scripts from LLM Spec"] = yara_llm_file
            row["LLM YARA # of lines"] = lines
            row["LLM YARA # of rules"] = rules
            row["LLM YARA # of Category"] = cats
            row["LLM YARA # of Sub-category"] = subs
        else:
            row["LLM Generated Yara Scripts from LLM Spec"] = "-"
            row["LLM YARA # of lines"] = 0
            row["LLM YARA # of rules"] = 0
            row["LLM YARA # of Category"] = 0
            row["LLM YARA # of Sub-category"] = 0

        data.append(row)

    df = pd.DataFrame(data)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = os.path.join(output_dir, f"Report_Summary_{timestamp}.xlsx")
    
    df.to_excel(output_file, index=False)

    # Formatting
    wb = load_workbook(output_file)
    ws = wb.active
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 45)

    wb.save(output_file)
    print(f"✅ Summary Report generated: {output_file}")


if __name__ == "__main__":
    generate_report_summary()